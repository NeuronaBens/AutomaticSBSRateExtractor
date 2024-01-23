from DateService import DateServiceImp
from Scrapper import WebScraperContable, WebScraperPromedio
import openpyxl

dateService = DateServiceImp()
scrapperContable = WebScraperContable("https://www.sbs.gob.pe/app/pp/SISTIP_PORTAL/Paginas/Publicacion/TipoCambioContable.aspx")
scrapperPromedio = WebScraperPromedio("https://www.sbs.gob.pe/app/pp/SISTIP_PORTAL/Paginas/Publicacion/TipoCambioPromedio.aspx")

correctDate = dateService.getCorrectDate()
contableDate = scrapperContable.scrape_date()
promedioDate = scrapperPromedio.scrape_date()

# Función para verificar si las fechas son correctas en la plataforma SBS.
# Compara las fechas correctas con las fechas en las variables contableDate y promedioDate.
# Si son iguales, devuelve True; de lo contrario, imprime un mensaje de error y devuelve False.
def verifyIfCorrectDates():
    if correctDate == contableDate and correctDate == promedioDate:
        return True
    else:
        print("incorrect dates on SBS platform")
        return False

# Función para actualizar celdas en una hoja de cálculo.
# Toma un número de fila, y una función de raspado (scrapeFunc) como argumentos.
# Extrae un valor usando la función de raspado, lo convierte a float y actualiza la celda en la columna 9.
# También actualiza la celda en la columna 3 con el valor de la variable correctDate.
def updateCells(row, scrapeFunc):
        column_index = 9
        cell_value = float(scrapeFunc())
        sheet.cell(row=row, column=column_index, value=cell_value)
        column_index = 3
        sheet.cell(row=row, column=column_index, value=correctDate)

if(verifyIfCorrectDates()):
    excel_file_path = 'Tabla de combinaciones TC - S4-SBS.xlsx'
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook['VF']

    #Dolar de N.A.
    updateCells(9, scrapperPromedio.scrape_dolar)

    # Dólar Australiano
    updateCells(10, scrapperContable.scrape_dolar_australiano)

    # Boliviano
    updateCells(11, scrapperContable.scrape_dolar_boliviano)

    # Dólar Canadiense
    updateCells(12, scrapperContable.scrape_dolar_canadiense)

    # Franco Suizo
    updateCells(13, scrapperContable.scrape_dolar_franco_suizo)

    # Peso Chileno
    updateCells(14, scrapperContable.scrape_dolar_peso_chileno)

    # Yuan
    updateCells(15, scrapperContable.scrape_dolar_yuan)

    # Peso Colombiano
    updateCells(16, scrapperContable.scrape_dolar_peso_colombiano)

    # Corona Danesa
    updateCells(17, scrapperContable.scrape_dolar_corona_danesa)

    # Euro
    updateCells(18, scrapperPromedio.scrape_euro)

    # Libra Esterlina
    updateCells(19, scrapperContable.scrape_dolar_libra_esterlina)

    # Quetzal
    updateCells(20, scrapperContable.scrape_dolar_quetzal)

    # Yen Japonés
    updateCells(21, scrapperContable.scrape_dolar_yen_japones)

    # Peso Mexicano
    updateCells(22, scrapperContable.scrape_dolar_peso_mexicano)

    # Corona Noruega
    updateCells(23, scrapperContable.scrape_dolar_corona_noruega)

    # Balboa
    updateCells(24, scrapperContable.scrape_dolar_balboa)

    # Corona Sueca
    updateCells(25, scrapperContable.scrape_dolar_corona_sueca)

    # Euro
    updateCells(26, scrapperPromedio.scrape_euro)


    workbook.save(excel_file_path)
    workbook.close()
